import numpy as np
import openpyxl
import xlwt
from scipy.optimize import linprog
import warnings
warnings.filterwarnings("ignore")
from sklearn.linear_model import LogisticRegression

LINES = 500
TEST = 10

wb = openpyxl.reader.excel.load_workbook(filename="sample.xlsx", data_only=True)
# print (wb.sheetnames)
wb.active = 0
sheet = wb.active


def sigmoid(x):
    return 1 / (1 + np.exp(-x))


array_in = [0] * (LINES + 1)
for i in range(1, LINES + 1):
    array_in[i] = [sheet['L' + str(i)].value, sheet['M' + str(i)].value, sheet['N' + str(i)].value,
                   sheet['O' + str(i)].value, sheet['P' + str(i)].value, sheet['Q' + str(i)].value]
array_in.remove(0)

training_inputs = np.array(array_in)
# print(array_in[499])
# print(training_inputs)


array_out = [0] * (LINES + 1)
my_arr = [0] * (LINES + 1)
for j in range(1, (LINES + 1)):
    newval = sheet['V' + str(j)].value
    array_out[j] = [newval]
    my_arr[j] = newval
array_out.remove(0)
my_arr.remove(0)


# training_outputs = np.array(array_out).T
training_outputs = np.array(array_out)
my_outputs = np.array(my_arr)
# print(training_outputs)


tests_num = 1
for m in range(tests_num):
    # for m in range (5000):
    # np.random.seed(m)

    #synaptic_weights = 2 * np.random.random((6, 1)) - 1

    #    print("Случайные инициализирующие веса:")
    # print(synaptic_weights)

    # ===============Задаем количество приближений
    #for i in range(100):
    #    input_layer = training_inputs
    #    outputs = sigmoid(np.dot(input_layer, synaptic_weights))

    #    err = training_outputs - outputs
    #    adjustments = np.dot(input_layer.T, err * (outputs * (1 - outputs)))

    #   synaptic_weights += adjustments

    LogisticRegression(penalty='l1', tol=0.01)
    X = np.linalg.lstsq(training_inputs, training_outputs, rcond=None)
    #linprog(training_outputs, training_inputs, training_outputs, training_inputs, training_outputs)
    print("==Веса после обучения:==")
    print(X[0])

    print("==Результат после обучения:==")
    #X = np.linalg.lstsq(training_inputs, training_outputs, rcond=None)
    outputs = np.dot(training_inputs, X[0])
    print(outputs)


    # создаем новый excel-файл
    wb = openpyxl.Workbook()

    # добавляем новый лист
    wb.create_sheet(title='Первый лист', index=0)

    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']

    for k in range(1, LINES + 1):
        result = outputs.flatten()
        sheet['S' + str(k)] = result[k - 1]
        # print(result[k-1])
        wb.save("example.xlsx")

    # Тест
    # new_inputs = np.array([0.56, 0.57, 0.74, 0.57, 0.41, 0.46])
    # new_inputs = np.array([0.79, 0.43, 1.0, 0.39, 0.40, 0.19])
    # print(new_inputs)
    #output = sigmoid(np.dot(new_inputs, synaptic_weights))
    #output = sigmoid(np.dot(new_inputs, synaptic_weights))

    print("==Берем данные из ВВОД ДАННЫХ.xlsx==")

    wb = openpyxl.reader.excel.load_workbook(filename="ВВОД ДАННЫХ.xlsx", data_only=True)
    wb.active = 0
    sheet = wb.active

    array_in = [0] * (TEST + 1)
    for i in range(1, TEST + 1):
        array_in[i] = [sheet['L' + str(i)].value, sheet['M' + str(i)].value, sheet['N' + str(i)].value,
                       sheet['O' + str(i)].value, sheet['P' + str(i)].value, sheet['Q' + str(i)].value]
    array_in.remove(0)
    new_inputs = np.array(array_in)

    # print("Введите 6 значений или 0 для завершения:")
    # line = input().split()
    # for i in range(len(line)):
    #    line[i] = float(line[i])
    # if not line[0]:
    #    break
    # new_inputs = np.array(line)

    res = []
    rres = []
    for i in range(TEST):
        try:
            linprog(new_inputs, training_inputs, training_outputs, training_inputs, training_outputs)
        except:
            pass
        for elem in X:
            if elem.size == 6:
                output = np.dot(new_inputs[i], elem)
                #print (output)
                # print(output, sigmoid(output))
                if output >= 0 and output <= 1:
                    #print(output)
                    routput = round(output[0] * 2) / 2
                    print("Result - ", output, " окр = ", routput)
                    rres.append(routput)
                    res.append(output[0])

    # wb = openpyxl.load_workbook("ВВОД ДАННЫХ.xlsx")
    # wb.active = 0
    # sheet = wb['Лист1']

    for k in range(1, TEST + 1):
        sheet['R' + str(k)] = res[k - 1]
        sheet['S' + str(k)] = rres[k - 1]
        # print(result[k-1])
    wb.save("ВВОД ДАННЫХ.xlsx")

    #X = np.linalg.lstsq(training_inputs, my_outputs, rcond=None)
    # print("Inp")
    # print(training_inputs)
    # print("Out")
    # print(my_outputs)
    #print("result - ", X)

print("=============================================================================================")
